import openpyxl

file_path = "student.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

midterm_weight = 0.3
final_weight = 0.35
hw_weight = 0.34
attendance_weight = 0.01

for row in sheet.iter_rows(min_row=2, values_only=True):
    midterm, final, hw, attendance = row[2], row[3], row[4], row[5]
    total_score = midterm * midterm_weight + final * final_weight + hw * hw_weight + attendance * attendance_weight

    sheet.cell(row=row[0], column=7, value=total_score)

workbook.save(file_path)
workbook.close()